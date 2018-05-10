import numpy as np
from openpyxl import load_workbook


def make_line(photo):
    title = photo[0]
    if len(title) > 57:
        title = title[:57] + "..."
    score_line = ""
    for i in range(len(photo[2])):
        score_line += str(photo[2][i]) + " | "

    score_line += str(photo[1])

    return "| %s | %s |" % (title, score_line)


def create_md_file(pics) -> None:
    print(pics)
    selected = pics[:10]
    last_entry_score = selected[-1][1]
    for i in range(len(selected), len(pics) - 1):
        if pics[i][1] == last_entry_score and last_entry_score != 0:
            selected.append(pics[i])

    header_line = "| Entry Title |"
    sub_header_line = "| --- | "
    for i in range(len(selected[0][2])):
        header_line += "Judge %d |" % (i + 1)
        sub_header_line += "--- | "

    header_line += " Total Score |"
    sub_header_line += "---|\n"
    with open('judge.md', "w") as o:
        header = "## Top %d Shortlisted entries for Judges' choice\n" % len(
            selected)
        print(header, file=o)
        print(header_line, file=o)
        print(sub_header_line, file=o)
        for s in selected:
            print(make_line(s), file=o)


def final_score(val):
    if val >= 80:
        return 5
    elif val >= 60:
        return 4
    elif val >= 40:
        return 3
    elif val >= 20:
        return 2
    else:
        return 1


def calculate_score(file_name):
    wb2 = load_workbook(file_name)
    ws = wb2['Sheet1']
    names = []
    name_and_score = []
    score = []
    for w in ws.iter_cols(0, 1):
        for c in w:
            names.append(c.value)

    for w in ws.iter_cols(2, ws.max_column):
        n1 = []
        for c in w:
            n1.append(c.value)
        score.append(n1)
    score = np.asarray(score)
    total = []
    for j in score:
        min_value = min(j[j > 0])
        max_value = max(j[j > 0])
        m = [(x - min_value) * 100 / (max_value - min_value) if (x > 0) else 0
             for x in j]
        for i in range(len(j)):
            if j[i] != 0:
                j[i] = final_score(m[i])

        total.append(j)
    total = np.asarray(total)
    for i in range(len(names)):
        name_and_score.append([names[i], sum(total[:, i]), total[:, i]])

    name_and_score = sorted(name_and_score, key=lambda l: l[1], reverse=True)
    create_md_file(name_and_score)
